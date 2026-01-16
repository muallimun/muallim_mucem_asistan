import sys, re, time, keyboard, pyperclip, mouse, json, os, traceback, urllib.request, ssl, ctypes, subprocess
import openpyxl
from openpyxl import Workbook, load_workbook
if getattr(sys, 'frozen', False):
    # Uygulamanın her zaman kendi kurulu olduğu klasörde (C:\Program Files vb.) çalışmasını sağlar
    os.chdir(os.path.dirname(sys.executable))

# --- GLOBAL DEĞİŞKENLER ---
VERSION = "1.5.0" #
GITHUB_VERSION_URL = "https://raw.githubusercontent.com/muallimun/muallim_mucem_asistan/refs/heads/main/version.json"

try: import winreg
except ImportError: winreg = None

from PyQt6.QtCore import Qt, QUrl, pyqtSignal, QObject, QTimer, QPoint
from PyQt6.QtWidgets import (QApplication, QMainWindow, QVBoxLayout, QWidget, 
                             QPushButton, QHBoxLayout, QLabel, QSystemTrayIcon, 
                             QMenu, QDialog, QLineEdit, QMessageBox, 
                             QCheckBox, QFileDialog, QFrame, QTextEdit, QScrollArea, QSizeGrip, QComboBox)
from PyQt6.QtWebEngineWidgets import QWebEngineView
from PyQt6.QtWebEngineCore import QWebEnginePage
from PyQt6.QtGui import QAction, QCursor, QDesktopServices, QIcon

# --- YÖNETİCİ KONTROLÜ VE SİSTEM FONKSİYONLARI ---

def is_admin():
    """Sistemin yönetici yetkisiyle çalışıp çalışmadığını kontrol eder."""
    try: return ctypes.windll.shell32.IsUserAnAdmin()
    except: return False

def run_as_admin():
    """Uygulamayı yönetici yetkisiyle (UAC) yeniden başlatır."""
    if is_admin(): return True
    executable = sys.executable
    if getattr(sys, 'frozen', False):
        params = ' '.join([f'"{arg}"' for arg in sys.argv[1:]])
    else:
        script_path = os.path.abspath(sys.argv[0])
        params = f'"{script_path}" ' + ' '.join([f'"{arg}"' for arg in sys.argv[1:]])
    
    try:
        ctypes.windll.shell32.ShellExecuteW(None, "runas", executable, params, None, 1)
    except: pass
    return False

def set_auto_start(enabled=True):
    """Windows Görev Zamanlayıcı kaydını en yüksek yetkiyle oluşturur."""
    task_name = "MuallimunAsistanAutoStart"
    
    # Kayıt defteri temizliği (Kodunuzdan korundu)
    if winreg:
        try:
            key = winreg.OpenKey(winreg.HKEY_CURRENT_USER, r"Software\Microsoft\Windows\CurrentVersion\Run", 0, winreg.KEY_SET_VALUE)
            winreg.DeleteValue(key, "MuallimunAsistan")
            winreg.CloseKey(key)
        except: pass

    app_path = os.path.abspath(sys.executable if getattr(sys, 'frozen', False) else sys.argv[0])
    
    # KRİTİK: Boşluklu yollar için tırnakları schtasks'ın anlayacağı şekilde kaçırıyoruz
    # /tr parametresi içindeki yol tırnak içine alınmalıdır: \"yol\" --parametre
    task_cmd = f'\\"{app_path}\\" --silent-start'

    try:
        # Mevcut görevi sil
        subprocess.run(f'schtasks /delete /tn "{task_name}" /f', shell=True, capture_output=True, creationflags=subprocess.CREATE_NO_WINDOW)
        
        if enabled:
            # GÖREVİ OLUŞTUR:
            # /rl highest: Yönetici onayı sormadan başlatır
            # /it: Kullanıcıyla etkileşime izin verir (Tray ikon için şart)
            # /np: Şifre sormamasını sağlar (Bazen tetikleme sorunlarını çözer)
            cmd = f'schtasks /create /tn "{task_name}" /tr "{task_cmd}" /sc onlogon /rl highest /it /f'
            subprocess.run(cmd, shell=True, capture_output=True, creationflags=subprocess.CREATE_NO_WINDOW)
    except: pass

def get_app_data_path():
    """Uygulama verileri ve loglar için klasör yolunu döner."""
    path = os.path.join(os.environ['APPDATA'], 'MuallimunAsistan')
    if not os.path.exists(path): os.makedirs(path)
    return path

def resource_path(relative_path):
    """Derlenmiş dosyada kaynak (ikon vb.) yollarını çözer."""
    try: base_path = sys._MEIPASS
    except Exception: base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)

# --- ÇOKLU DİL SİSTEMİ ---
TRANSLATIONS = {
    "tr": {
        "app_title": "Muallimun Sözlük v",
        "settings_title": "Muallimun Asistan Ayarları",
        "guide_title": "🚀 Detaylı Kullanım Rehberi",
        "guide_text": "<b>• Evrensel Erişim:</b> Her türlü pencerede metin yakalar.<br><b>• PDF ve Belgeler:</b> Kelimeyi seçip <b>Mouse Orta Tekerlek</b> tıklayın.<br><b>• Web Sayfaları:</b> Kelimeyi seçip <b>{}</b> tuşlayın.<br><b>• Akıllı Algılama:</b> Dil uyumsuzluğunu sistem otomatik giderir.<br><b>• Kayıt:</b> Aramalar Excel dosyanıza tarihli işlenir.",
        "update_btn": "🔄 Güncelleme Kontrol Et",
        "dict_mode": "📚 Varsayılan Sözlük Modu:",
        "lang_label": "🌐 Uygulama Dili:",
        "hotkey_label": "⌨️ Global Kısayol Tuşu:",
        "excel_label": "📂 Excel Kayıt Dosyası Yolu:",
        "browse": "Gözat...",
        "auto_start": "Bilgisayar açılışında otomatik başlat",
        "startup_warning": "⚠️ Uygulamayı kurduktan sonra bu seçeneği bir kez kapatıp tekrar açın. Bu, Windows kayıtlarını güncelleyecektir.", #
        "empty_save": "Anlam girilmeden kayda izin ver",
        "save_apply": "Ayarları Kaydet ve Uygula",
        "online_guide": "Muallimun.Net Online Rehber",
        "meaning_placeholder": "Anlamı buraya yazın veya tarayıcıdan sürükleyin...",
        "empty_defter": "Kelime anlamı olmadan deftere kaydet",
        "excel_save_btn": "Excel'e Kaydet",
        "excel_success": "Kaydedildi ✅",
        "excel_busy": "Excel dosyası açık! Lütfen kapatıp tekrar deneyin.",
        "ready_msg": "Arka planda hazır!",
        "tray_settings": "Ayarlar",
        "tray_exit": "Çıkış",
        "tt_settings": "Ayarları Aç",
        "tt_close": "Kapat",
        "tt_ontop": "Pencereyi Üstte Tut / Sabitle",
        "tt_browser": "Tarayıcıda Tam Sayfa Aç",
        "upd_title": "Güncelleme",
        "upd_new": "Yeni bir sürüm mevcut. İndirme sayfasına gitmek ister misiniz?",
        "upd_info": "Bilgi",
        "upd_latest": "Uygulama güncel. En son sürümü kullanıyorsunuz.",
        "modes": ["Arapça <> Türkçe", "Arapça <> Arapça", "Arapça <> İngilizce"]
    },
    "en": {
        "app_title": "Muallimun Dictionary v",
        "settings_title": "Muallimun Assistant Settings",
        "guide_title": "🚀 Detailed User Guide",
        "guide_text": "<b>• Universal Access:</b> Captures text in all windows.<br><b>• PDF & Docs:</b> Select word and click <b>Mouse Middle Wheel</b>.<br><b>• Web Pages:</b> Select word and press <b>{}</b>.<br><b>• Smart Detection:</b> System handles language mismatches.<br><b>• Save:</b> Logs searches with dates into Excel.",
        "update_btn": "🔄 Check for Updates",
        "dict_mode": "📚 Default Dictionary Mode:",
        "lang_label": "🌐 Application Language:",
        "hotkey_label": "⌨️ Global Hotkey:",
        "excel_label": "📂 Excel Log File Path:",
        "browse": "Browse...",
        "auto_start": "Start automatically on Windows login",
        "startup_warning": "⚠️ After installation, please toggle this option off and on once to update Windows task records.", #
        "empty_save": "Allow saving without meaning",
        "save_apply": "Save and Apply Settings",
        "online_guide": "Muallimun.Net Online Guide",
        "meaning_placeholder": "Type meaning here or drag from browser...",
        "empty_defter": "Save without meaning",
        "excel_save_btn": "Save to Excel",
        "excel_success": "Saved ✅",
        "excel_busy": "Excel busy! Please close it.",
        "ready_msg": "Ready in background!",
        "tray_settings": "Settings",
        "tray_exit": "Exit",
        "tt_settings": "Open Settings",
        "tt_close": "Close",
        "tt_ontop": "Stay on Top / Pin Window",
        "tt_browser": "Open in Full Browser",
        "upd_title": "Update",
        "upd_new": "A new version is available. Do you want to go to the download page?",
        "upd_info": "Information",
        "upd_latest": "App is up to date. You are using the latest version.",
        "modes": ["Arabic <> Turkish", "Arabic <> Arabic", "Arabic <> English"]
    },
    "ar": {
        "app_title": "مساعد معجم معلمون v",
        "settings_title": "إعدادات مساعد معلمون",
        "guide_title": "🚀 دليل الاستخدام المفصل",
        "guide_text": "<b>• الوصول الشامل:</b> يلتقط النصوص في جميع أنواع النوافذ.<br><b>• ملفات PDF والمستندات:</b> اختر الكلمة ثم اضغط على <b>زر الفأرة الأوسط</b>.<br><b>• صفحات الويب:</b> اختر الكلمة ثم اضغط على <b>{}</b>.<br><b>• الكشف الذكي:</b> يقوم النظام بمعالجة عدم تطابق اللغة تلقائيًا.<br><b>• السجل الذكي:</b> يتم تسجيل عمليات البحث مع التواريخ في ملف إكسل.",
        "update_btn": "🔄 التحقق من التحديثات",
        "dict_mode": "📚 وضع القاموس الافتراضي:",
        "lang_label": "🌐 لغة التطبيق:",
        "hotkey_label": "⌨️ مفتاح الاختصار العالمي:",
        "excel_label": "📂 مسار ملف إكسل للسجل:",
        "browse": "تصفح...",
        "auto_start": "التشغيل تلقائيًا عند بدء ويندوز",
        "startup_warning": "⚠️ بعد التثبيت، يرجى إيقاف تشغيل هذا الخيار ثم تشغيله مرة أخرى لتحديث سجلات ويندوز.", #
        "empty_save": "السماح بالحفظ بدون معنى",
        "save_apply": "حفظ وتطبيق الإعدادات",
        "online_guide": "دليل معلمون.نت على الإنترنت",
        "meaning_placeholder": "اكتب المعنى هنا أو اسحبه من المتصفح...",
        "empty_defter": "الحفظ في الدفتر بدون معنى",
        "excel_save_btn": "حفظ في إكسل",
        "excel_success": "تم الحفظ ✅",
        "excel_busy": "ملف إكسل مفتوح! يرجى إغلاقه والمحاولة مرة أخرى.",
        "ready_msg": "جاهز في الخلفية!",
        "tray_settings": "الإعدادات",
        "tray_exit": "خروج",
        "tt_settings": "فتح إعدادات التطبيق",
        "tt_close": "إغلاق النافذة",
        "tt_ontop": "تثبيت النافذة في الأعلى",
        "tt_browser": "فتح في المتصفح بالكامل",
        "upd_title": "تحديث",
        "upd_new": "يوجد إصدار جديد متاح. هل تريد الذهاب إلى صفحة التحميل؟",
        "upd_info": "معلومات",
        "upd_latest": "التطبيق محدث. أنت تستخدم أحدث إصدار.",
        "modes": ["عربي <> تركي", "عربي <> عربي", "عربي <> إنجليزي"]
    }
}

class SettingsManager:
    def __init__(self):
        self.path = os.path.join(get_app_data_path(), "asistan_ayarlar.json")
        # lang: "en" yapıldı ve varsayılan sözlük modu İngilizce (index: 2) olarak ayarlandı
        self.defaults = {"hotkey": "ctrl+shift+z", "excel_path": os.path.join(os.path.expanduser("~"), "Desktop", "Arabic_Word_Bank.xlsx"), "auto_start": False, "allow_empty_meaning": False, "dict_mode_index": 2, "lang": "en"}
        if not os.path.exists(self.path): self.save(self.defaults)
    def load(self):
        try:
            with open(self.path, "r", encoding="utf-8") as f: return json.load(f)
        except: return self.defaults
    def save(self, data):
        with open(self.path, "w", encoding="utf-8") as f: json.dump(data, f, indent=4)
        set_auto_start(data.get("auto_start", False))

class SilentWebPage(QWebEnginePage):
    def javaScriptConsoleMessage(self, level, message, lineID, sourceID): pass

class SettingsDialog(QDialog):
    settings_changed = pyqtSignal()
    def __init__(self, manager, parent=None):
        super().__init__(parent); self.manager = manager; self.settings = manager.load()
        self.lang = self.settings.get("lang", "tr"); self.t = TRANSLATIONS[self.lang]
        self.setWindowTitle(self.t["settings_title"]); self.setFixedSize(560, 720)
        self.setStyleSheet("background-color: white;"); self.setWindowIcon(QIcon(resource_path("muallim.ico")))
        layout = QVBoxLayout(self); layout.setSpacing(1); layout.setContentsMargins(12, 12, 12, 12)
        
        info_frame = QFrame(); info_frame.setStyleSheet("background-color: #f8fbff; border-radius: 12px; border: 1px solid #e2e8f0;")
        info_layout = QVBoxLayout(info_frame); info_layout.setContentsMargins(15, 6, 15, 6); info_layout.setSpacing(1)
        
        # Güncellenecek etiketleri self. ile tanımlıyoruz
        self.guide_title = QLabel(self.t["guide_title"]); self.guide_title.setStyleSheet("font-size: 16px; font-weight: bold; color: #1e3a8a;")
        self.guide_text_lbl = QLabel(self.t["guide_text"].format(self.settings['hotkey'].upper()))
        self.guide_text_lbl.setWordWrap(True); self.guide_text_lbl.setStyleSheet("color: #334155; line-height: 115%; font-size: 11.5px;")
        self.btn_update = QPushButton(self.t["update_btn"]); self.btn_update.setStyleSheet("background: #3b82f6; color: white; font-weight: bold; padding: 6px; border-radius: 6px; border:none;")
        self.btn_update.clicked.connect(lambda: self.check_update(manual=True))
        info_layout.addWidget(self.guide_title); info_layout.addWidget(self.guide_text_lbl); info_layout.addWidget(self.btn_update)
        layout.addWidget(info_frame)
        
        form_frame = QFrame(); form_frame.setStyleSheet("background: white; border: 1px solid #e2e8f0; border-radius: 10px; padding: 6px;")
        form_layout = QVBoxLayout(form_frame); form_layout.setSpacing(3)
        
        self.lbl_lang = QLabel(f"<b>{self.t['lang_label']}</b>"); form_layout.addWidget(self.lbl_lang)
        self.lang_combo = QComboBox(); self.lang_combo.addItems(["Türkçe", "English", "العربية"])
        self.lang_combo.setCurrentText("Türkçe" if self.lang == "tr" else "English" if self.lang == "en" else "العربية")
        form_layout.addWidget(self.lang_combo)
        
        self.lbl_mode = QLabel(f"<b>{self.t['dict_mode']}</b>"); form_layout.addWidget(self.lbl_mode)
        self.mode_combo = QComboBox(); self.mode_combo.addItems(self.t["modes"]); self.mode_combo.setCurrentIndex(self.settings.get("dict_mode_index", 0)); form_layout.addWidget(self.mode_combo)
        
        self.lbl_hotkey = QLabel(f"<b>{self.t['hotkey_label']}</b>"); form_layout.addWidget(self.lbl_hotkey)
        self.hotkey_input = QLineEdit(self.settings["hotkey"]); self.hotkey_input.setStyleSheet("padding: 4px; border: 1px solid #cbd5e1; border-radius: 4px;"); form_layout.addWidget(self.hotkey_input)
        
        self.lbl_excel = QLabel(f"<b>{self.t['excel_label']}</b>"); form_layout.addWidget(self.lbl_excel)
        p_lay = QHBoxLayout(); self.path_input = QLineEdit(self.settings["excel_path"]); self.path_input.setStyleSheet("padding: 4px; border: 1px solid #cbd5e1; border-radius: 4px;")
        self.btn_browse = QPushButton(self.t["browse"]); self.btn_browse.clicked.connect(self.browse_path); p_lay.addWidget(self.path_input); p_lay.addWidget(self.btn_browse); form_layout.addLayout(p_lay)
        
        cb_style = "QCheckBox { color: #334155; font-size: 11.5px; } QCheckBox::indicator { width: 15px; height: 15px; }"
        self.auto_start_cb = QCheckBox(self.t["auto_start"]); self.auto_start_cb.setStyleSheet(cb_style); self.auto_start_cb.setChecked(self.settings.get("auto_start", False))
        form_layout.addWidget(self.auto_start_cb)
        
        self.startup_note = QLabel(self.t["startup_warning"])
        self.startup_note.setWordWrap(True); self.startup_note.setStyleSheet("color: #e67e22; font-size: 10px; font-style: italic; margin-left: 20px; margin-bottom: 5px;")
        form_layout.addWidget(self.startup_note)
        
        self.empty_save_cb = QCheckBox(self.t["empty_save"]); self.empty_save_cb.setStyleSheet(cb_style); self.empty_save_cb.setChecked(self.settings.get("allow_empty_meaning", False))
        form_layout.addWidget(self.empty_save_cb); layout.addWidget(form_frame)
        
        self.btn_save_settings = QPushButton(self.t["save_apply"]); self.btn_save_settings.setFixedHeight(45); self.btn_save_settings.setStyleSheet("background: #2ecc71; color: white; font-weight: bold; border-radius: 8px; font-size: 14px; border:none;")
        self.btn_save_settings.clicked.connect(self.save_settings); layout.addWidget(self.btn_save_settings)
        
        footer_layout = QHBoxLayout()
        self.f_link = QLabel(f'<a href="https://arapca.muallimun.net/asistan_sozluk/" style="color: #1e3a8a; text-decoration: none; font-weight: bold;">{self.t["online_guide"]}</a>')
        self.f_link.setOpenExternalLinks(True)
        self.version_lbl = QLabel(f"v{VERSION}"); self.version_lbl.setStyleSheet("color: #94a3b8; font-size: 11px; font-weight: normal;")
        footer_layout.addWidget(self.f_link); footer_layout.addStretch(); footer_layout.addWidget(self.version_lbl); layout.addLayout(footer_layout)

    def browse_path(self):
        f, _ = QFileDialog.getSaveFileName(self, self.t["browse"], self.path_input.text(), "Excel Files (*.xlsx)")
        if f: self.path_input.setText(f)

    def check_update(self, manual=False):
        try:
            ctx = ssl._create_unverified_context(); req = urllib.request.Request(GITHUB_VERSION_URL + f"?t={int(time.time())}")
            with urllib.request.urlopen(req, context=ctx) as r:
                data = json.loads(r.read().decode('utf-8'))
                if str(data.get("version")).strip() != VERSION:
                    # Dile duyarlı yeni sürüm uyarısı
                    if QMessageBox.information(self, self.t["upd_title"], self.t["upd_new"], 
                                               QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No) == QMessageBox.StandardButton.Yes:
                        QDesktopServices.openUrl(QUrl(data.get("url")))
                elif manual: 
                    # Dile duyarlı "güncel" mesajı
                    QMessageBox.information(self, self.t["upd_info"], self.t["upd_latest"])
        except: pass

    def retranslate_ui(self):
        """Pencereyi kapatmadan tüm metinleri yeni dile göre günceller."""
        self.t = TRANSLATIONS[self.lang]
        self.setWindowTitle(self.t["settings_title"])
        self.guide_title.setText(self.t["guide_title"])
        self.guide_text_lbl.setText(self.t["guide_text"].format(self.settings['hotkey'].upper()))
        self.btn_update.setText(self.t["update_btn"])
        self.lbl_lang.setText(f"<b>{self.t['lang_label']}</b>")
        self.lbl_mode.setText(f"<b>{self.t['dict_mode']}</b>")
        self.lbl_hotkey.setText(f"<b>{self.t['hotkey_label']}</b>")
        self.lbl_excel.setText(f"<b>{self.t['excel_label']}</b>")
        self.btn_browse.setText(self.t["browse"])
        self.auto_start_cb.setText(self.t["auto_start"])
        self.startup_note.setText(self.t["startup_warning"])
        self.empty_save_cb.setText(self.t["empty_save"])
        self.btn_save_settings.setText(self.t["save_apply"])
        self.f_link.setText(f'<a href="https://arapca.muallimun.net/asistan_sozluk/" style="color: #1e3a8a; text-decoration: none; font-weight: bold;">{self.t["online_guide"]}</a>')
        
        # Sözlük modlarını dile göre yenile (indeksi koruyarak)
        curr_idx = self.mode_combo.currentIndex()
        self.mode_combo.clear(); self.mode_combo.addItems(self.t["modes"]); self.mode_combo.setCurrentIndex(curr_idx)
        
        # Butona geçici 'Kaydedildi' onayı ver
        old_text = self.t["save_apply"]
        self.btn_save_settings.setText("✅ " + self.t["save_apply"])
        QTimer.singleShot(1500, lambda: self.btn_save_settings.setText(old_text))

    def save_settings(self):
        lt = self.lang_combo.currentText(); nl = "tr" if lt == "Türkçe" else "en" if lt == "English" else "ar"
        self.lang = nl # Sınıfın aktif dilini güncelle
        self.settings.update({"hotkey": self.hotkey_input.text().lower().strip(), "excel_path": self.path_input.text().strip(), "auto_start": self.auto_start_cb.isChecked(), "allow_empty_meaning": self.empty_save_cb.isChecked(), "dict_mode_index": self.mode_combo.currentIndex(), "lang": nl})
        self.manager.save(self.settings); self.settings_changed.emit()
        self.retranslate_ui() # Pencereyi kapatmadan metinleri güncelle

class DictionaryWindow(QMainWindow):
    open_settings_signal = pyqtSignal()
    def __init__(self, settings):
        super().__init__(); self.settings = settings
        self.lang = self.settings.get("lang", "tr"); self.t = TRANSLATIONS[self.lang]
        self.setWindowFlags(Qt.WindowType.WindowStaysOnTopHint | Qt.WindowType.Window | Qt.WindowType.CustomizeWindowHint)
        self.setWindowTitle(f"{self.t['app_title']}{VERSION}"); self.setMinimumSize(600, 500); self.resize(620, 750)
        self.setWindowIcon(QIcon(resource_path("muallim.ico"))); self._old_pos = None
        c = QWidget(); c.setStyleSheet("QWidget { background: white; border: 1px solid #1e3a8a; border-radius: 12px; }")
        layout = QVBoxLayout(c); layout.setContentsMargins(0, 0, 0, 0)
        
        # --- HEADER (ÜST ÇUBUK) ---
        self.header = QWidget(); self.header.setFixedHeight(45); self.header.setStyleSheet("background: #1e3a8a; border-top-left-radius: 10px; border-top-right-radius: 10px; border:none;")
        h_lay = QHBoxLayout(self.header)
        self.title_lbl = QLabel(f"{self.t['app_title']}{VERSION}"); self.title_lbl.setStyleSheet("color: white; font-weight: bold; font-size: 14px;")
        
        # Buton Tanımlamaları
        self.btn_set = QPushButton("⚙"); self.btn_set.setFixedSize(30,30); self.btn_set.setStyleSheet("background: #64748b; color: white; border-radius: 15px;")
        self.btn_set.clicked.connect(self.open_settings_signal.emit)
        
        self.btn_ontop = QPushButton("📌"); self.btn_ontop.setCheckable(True); self.btn_ontop.setChecked(True)
        self.btn_ontop.setFixedSize(30, 30); self.btn_ontop.setStyleSheet("background: #10b981; color: white; border-radius: 15px; border:none;")
        self.btn_ontop.clicked.connect(self.toggle_on_top)
        
        self.btn_browser = QPushButton("🌍"); self.btn_browser.setFixedSize(30, 30); self.btn_browser.setStyleSheet("background: #64748b; color: white; border-radius: 15px; border:none;")
        self.btn_browser.clicked.connect(self.open_in_full_browser)
        
        self.btn_cls = QPushButton("✕"); self.btn_cls.setFixedSize(30,30); self.btn_cls.setStyleSheet("background: #ef4444; color: white; border-radius: 15px;")
        self.btn_cls.clicked.connect(self.hide)

        # Butonları Layout'a Ekleme
        h_lay.addWidget(self.title_lbl); h_lay.addStretch()
        h_lay.addWidget(self.btn_browser); h_lay.addWidget(self.btn_ontop); h_lay.addWidget(self.btn_set); h_lay.addWidget(self.btn_cls)
        
        # --- BROWSER VE ALT PANEL ---
        self.browser = QWebEngineView(); self.browser.setPage(SilentWebPage(self.browser)); self.browser.loadFinished.connect(self.clean_web)
        
        f = QWidget(); f.setFixedHeight(230); f.setStyleSheet("background: #f8fafc; border-top: 1px solid #e2e8f0; border:none;")
        f_lay = QVBoxLayout(f)
        
        self.meaning_box = QTextEdit(); self.meaning_box.setPlaceholderText(self.t["meaning_placeholder"])
        self.meaning_box.setStyleSheet("background: white; border: 1px solid #cbd5e1; border-radius: 8px; padding: 10px;")
        
        br = QHBoxLayout()
        self.quick_cb = QCheckBox(self.t["empty_defter"]); self.quick_cb.setStyleSheet("color: #334155; font-size: 13px;")
        
        self.btn_save = QPushButton(self.t["excel_save_btn"]); self.btn_save.setFixedHeight(40)
        self.btn_save.setStyleSheet("background: #f59e0b; color: white; font-weight: bold; border-radius: 10px;")
        self.btn_save.clicked.connect(self.save_to_excel)
        
        br.addWidget(self.quick_cb); br.addWidget(self.btn_save)
        
        f_lay.addWidget(QLabel("💡 <b>Meaning:</b>")); f_lay.addWidget(self.meaning_box); f_lay.addLayout(br)
        
        bot = QHBoxLayout()
        self.guide_lbl = QLabel(f'<a href="https://arapca.muallimun.net/asistan_sozluk/" style="color: #1e3a8a; text-decoration: none; font-weight: bold;">{self.t["online_guide"]}</a>')
        self.guide_lbl.setOpenExternalLinks(True)
        bot.addWidget(self.guide_lbl); bot.addStretch(); bot.addWidget(QLabel(f"v{VERSION}"))
        
        f_lay.addLayout(bot); layout.addWidget(self.header); layout.addWidget(self.browser); layout.addWidget(f)
        
        self.setCentralWidget(c); self.grip = QSizeGrip(self); layout.addWidget(self.grip, 0, Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignBottom)
        
        self.apply_tooltips()

    # --- YENİ FONKSİYONLAR ---
    def toggle_on_top(self):
        """Pencerenin her zaman üstte kalma durumunu değiştirir."""
        is_on = self.btn_ontop.isChecked()
        if is_on:
            self.setWindowFlags(self.windowFlags() | Qt.WindowType.WindowStaysOnTopHint)
            self.btn_ontop.setStyleSheet("background: #10b981; color: white; border-radius: 15px; border:none;")
        else:
            self.setWindowFlags(self.windowFlags() & ~Qt.WindowType.WindowStaysOnTopHint)
            self.btn_ontop.setStyleSheet("background: #64748b; color: white; border-radius: 15px; border:none;")
        self.show()

    def open_in_full_browser(self):
        """Mevcut kelimeyi varsayılan web tarayıcısında açar."""
        if hasattr(self, 'browser') and not self.browser.url().isEmpty():
            QDesktopServices.openUrl(self.browser.url())

    def apply_tooltips(self):
        """Tüm bileşenlerin ipucu metinlerini dile göre günceller."""
        self.btn_set.setToolTip(self.t["tt_settings"])
        self.btn_cls.setToolTip(self.t["tt_close"])
        self.btn_save.setToolTip(self.t["excel_save_btn"])
        self.quick_cb.setToolTip(self.t["empty_defter"])
        self.btn_ontop.setToolTip(self.t["tt_ontop"])
        self.btn_browser.setToolTip(self.t["tt_browser"])

    def clean_web(self):
        js = "var h=['header', 'footer', 'nav', '.header', '.ads', '.ad-unit', '.side-bar', '#top-nav', 'iframe', 'ins']; function hideAds() { h.forEach(s => document.querySelectorAll(s).forEach(n => n.style.display = 'none')); } hideAds(); window.scrollTo(0,0); setInterval(hideAds, 3000);"
        self.browser.page().runJavaScript(js)

    def update_texts(self, s):
        self.settings = s; self.lang = s.get("lang", "tr"); self.t = TRANSLATIONS[self.lang]
        self.setWindowTitle(f"{self.t['app_title']}{VERSION}"); self.title_lbl.setText(f"{self.t['app_title']}{VERSION}")
        self.meaning_box.setPlaceholderText(self.t["meaning_placeholder"]); self.quick_cb.setText(self.t["empty_defter"])
        self.btn_save.setText(self.t["excel_save_btn"]); self.guide_lbl.setText(f'<a href="https://arapca.muallimun.net/asistan_sozluk/" style="color: #1e3a8a; text-decoration: none; font-weight: bold;">{self.t["online_guide"]}</a>'); self.apply_tooltips()

    def search_word(self, word):
        self.current_word = word; self.meaning_box.clear()
        m_idx = self.settings.get("dict_mode_index", 0)
        
        # Site arayüz dilleri ve sözlük kodları eşleştirmesi
        # 0: Türkçe Site / ar-tr
        # 1: Arapça Site / ar-ar
        # 2: İngilizce Site / ar-en
        dict_langs = ["tr", "ar", "en"]
        dict_codes = ["ar-tr", "ar-ar", "ar-en"]
        
        # Seçilen indekse göre URL'yi doğrudan oluştur (Almaany çift yönlü çalışır)
        url = f"https://www.almaany.com/{dict_langs[m_idx]}/dict/{dict_codes[m_idx]}/{word}"
        
        self.browser.setUrl(QUrl(url)); self.showNormal(); self.show(); self.activateWindow()

    def save_to_excel(self):
        f = self.settings["excel_path"]; m = self.meaning_box.toPlainText().strip()
        if not m and not (self.settings.get("allow_empty_meaning") or self.quick_cb.isChecked()): return
        try:
            if not os.path.exists(f): wb = Workbook(); ws = wb.active; ws.append(["Date", "Word", "Meaning"])
            else: wb = load_workbook(f); ws = wb.active
            ws.append([time.strftime("%d.%m.%Y %H:%M"), getattr(self, "current_word", "---"), m or "---"]); wb.save(f)
            self.btn_save.setText(self.t["excel_success"]); QTimer.singleShot(2000, lambda: self.btn_save.setText(self.t["excel_save_btn"])); self.meaning_box.clear()
        except: QMessageBox.critical(self, "Error", self.t["excel_busy"])

    def mousePressEvent(self, e):
        if e.button() == Qt.MouseButton.LeftButton and self.header.underMouse(): self._old_pos = e.globalPosition().toPoint()
    def mouseMoveEvent(self, e):
        if self._old_pos: d = e.globalPosition().toPoint() - self._old_pos; self.move(self.x() + d.x(), self.y() + d.y()); self._old_pos = e.globalPosition().toPoint()
    def mouseReleaseEvent(self, e): self._old_pos = None

class AppController(QObject):
    search_signal = pyqtSignal(str)
    def __init__(self, start_minimized=False):
        super().__init__(); self.m = SettingsManager(); self.s = self.m.load(); self.t = TRANSLATIONS[self.s.get("lang", "tr")]
        self.app = QApplication(sys.argv); self.app.setQuitOnLastWindowClosed(False)
        self.win = DictionaryWindow(self.s); self.search_signal.connect(self.win.search_word); self.win.open_settings_signal.connect(self.open_settings)
        QTimer.singleShot(1000, self.setup_tray); self.refresh_listeners(); QTimer.singleShot(10000, self.auto_update_check)

    def setup_tray(self):
        try:
            self.tray = QSystemTrayIcon(QIcon(resource_path("muallim.ico")))
            self.tm = QMenu(); self.set_tm(); self.tray.setContextMenu(self.tm); self.tray.show()
            self.tray.showMessage("Muallimun", self.t["ready_msg"], QSystemTrayIcon.MessageIcon.Information, 3000)
        except: pass

    def set_tm(self):
        self.tm.clear(); self.tm.addAction(self.t["tray_settings"]).triggered.connect(self.open_settings); self.tm.addAction(self.t["tray_exit"], self.shutdown)

    def open_settings(self):
        d = SettingsDialog(self.m, self.win); d.settings_changed.connect(self.refresh_listeners); d.exec()

    def auto_update_check(self):
        """Arka planda sessizce güncelleme kontrolü yapar ve dile göre bildirim gönderir."""
        try:
            ctx = ssl._create_unverified_context(); req = urllib.request.Request(GITHUB_VERSION_URL + f"?t={int(time.time())}")
            with urllib.request.urlopen(req, context=ctx) as r:
                data = json.loads(r.read().decode('utf-8'))
                if str(data.get("version")).strip() != VERSION:
                    # Sabit metinler yerine TRANSLATIONS sözlüğündeki karşılıklar kullanıldı
                    self.tray.showMessage(
                        self.t["upd_title"], 
                        self.t["upd_new"], 
                        QSystemTrayIcon.MessageIcon.Information,
                        5000 # Bildirim 5 saniye görünür kalır
                    )
        except: pass

    def refresh_listeners(self):
        self.s = self.m.load(); self.t = TRANSLATIONS[self.s.get("lang", "tr")]; self.win.update_texts(self.s)
        try: self.set_tm(); keyboard.unhook_all_hotkeys(); mouse.unhook_all()
        except: pass
        keyboard.add_hotkey(self.s.get("hotkey", "ctrl+shift+z"), lambda: QTimer.singleShot(100, self.process), suppress=True)
        def cb(e):
            if isinstance(e, mouse.ButtonEvent) and e.button == mouse.MIDDLE and e.event_type == mouse.DOWN: QTimer.singleShot(10, self.process); return False
            return True
        mouse.hook(cb)

    def process(self):
        pyperclip.copy(""); keyboard.press_and_release('ctrl+c'); time.sleep(0.4); t = pyperclip.paste().strip()
        c = re.sub(r'[^\u0600-\u06FFa-zA-ZğüşıöçĞÜŞİÖÇİı\s]', '', t).strip()
        if c: self.search_signal.emit(c)

    def shutdown(self):
        try: keyboard.unhook_all_hotkeys(); mouse.unhook_all()
        except: pass
        self.app.quit()

    def run(self): return self.app.exec()

# ============================================================================
# ANA BAŞLATMA MANTIĞI (TAMAMEN REVİZE EDİLDİ)
# ============================================================================

if __name__ == "__main__":
    is_silent_start = "--silent-start" in sys.argv
    
    # 1. YÖNETİCİ KONTROLÜ
    # Silent start modunda Task Scheduler zaten en yüksek yetkiyi (Highest) sağlar.
    if not is_silent_start and not is_admin():
        # Manuel başlatmada tek seferlik UAC onayı iste ve eski süreci kapat
        run_as_admin()
        sys.exit(0)

    # 2. MUTEX KONTROLÜ (YÖNETİCİ OLDUKTAN SONRA)
    # Global takısı kullanarak farklı bütünlük seviyelerindeki süreçlerin çakışmasını engelliyoruz.
    k32 = ctypes.windll.kernel32
    m_name = f"Global\\MuallimunAsistan_v{VERSION.replace('.', '_')}_Final" # İsim çakışmasını önlemek için sabitlendi
    mutex = k32.CreateMutexW(None, False, m_name)
    
    if k32.GetLastError() == 183: # ERROR_ALREADY_EXISTS
        sys.exit(0)

    # 3. UYGULAMAYI BAŞLAT
    try:
        c = AppController(start_minimized=is_silent_start)
        sys.exit(c.run())
    except Exception as e:
        # Hataları teknik detaylı olarak crash_log.txt dosyasına yaz
        log_path = os.path.join(get_app_data_path(), "crash_log.txt")
        with open(log_path, "a", encoding="utf-8") as f:
            f.write(f"\n--- {time.ctime()} ---\n{traceback.format_exc()}\n")
        sys.exit(1)